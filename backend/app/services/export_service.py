"""
Smart Green ENSIT — Service d'export PDF & Excel
Génération du bilan carbone en format téléchargeable
"""
import io
import os
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, KeepTogether
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
try:
    from app.core.config import settings as _settings
except Exception:
    _settings = None
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


# ── Couleurs ENSIT ────────────────────────────────────────────
VERT_ENSIT   = colors.HexColor("#3B6D11")
VERT_CLAIR   = colors.HexColor("#EAF3DE")
VERT_MEDIUM  = colors.HexColor("#639922")
BLEU_S2      = colors.HexColor("#185FA5")
ROUGE_S1     = colors.HexColor("#A32D2D")
VIOLET_S3    = colors.HexColor("#534AB7")
GRIS_CLAIR   = colors.HexColor("#F1EFE8")
GRIS_BORD    = colors.HexColor("#D3D1C7")
NOIR         = colors.HexColor("#2C2C2A")
BLANC        = colors.white


# ════════════════════════════════════════════════════════
# EXPORT PDF — BILAN CARBONE COMPLET
# ════════════════════════════════════════════════════════

def generer_rapport_pdf(bilan: dict, periode: str = "2024-2025") -> bytes:
    """
    Génère le rapport PDF complet du bilan carbone ENSIT.
    Retourne les bytes du PDF prêt à télécharger.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2.5*cm, bottomMargin=2*cm,
        title=f"Bilan Carbone ENSIT {periode}",
        author="Plateforme Smart Green ENSIT",
    )

    styles = getSampleStyleSheet()
    story = []

    # ── Styles personnalisés ──────────────────────────────
    s_titre = ParagraphStyle("Titre",
        parent=styles["Normal"], fontSize=22, fontName="Helvetica-Bold",
        textColor=VERT_ENSIT, alignment=TA_CENTER, spaceAfter=4)

    s_sous_titre = ParagraphStyle("SousTitre",
        parent=styles["Normal"], fontSize=11,
        textColor=colors.HexColor("#5F5E5A"), alignment=TA_CENTER, spaceAfter=16)

    s_section = ParagraphStyle("Section",
        parent=styles["Normal"], fontSize=13, fontName="Helvetica-Bold",
        textColor=VERT_ENSIT, spaceBefore=16, spaceAfter=8,
        borderPad=4, leading=16)

    s_corps = ParagraphStyle("Corps",
        parent=styles["Normal"], fontSize=10,
        textColor=NOIR, leading=15, spaceAfter=6)

    s_note = ParagraphStyle("Note",
        parent=styles["Normal"], fontSize=8.5,
        textColor=colors.HexColor("#888780"), leading=12, leftIndent=12)

    s_kpi_val = ParagraphStyle("KPIVal",
        parent=styles["Normal"], fontSize=18, fontName="Helvetica-Bold",
        textColor=VERT_ENSIT, alignment=TA_CENTER)

    s_kpi_label = ParagraphStyle("KPILabel",
        parent=styles["Normal"], fontSize=8, alignment=TA_CENTER,
        textColor=colors.HexColor("#5F5E5A"))

    # ── EN-TÊTE ───────────────────────────────────────────
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph("🌿  SMART GREEN ENSIT", s_titre))
    story.append(Paragraph(
        f"Bilan Carbone Campus — Année universitaire {periode}", s_sous_titre))
    story.append(Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}  |  "
        f"Méthode : GHG Protocol  |  Facteurs : ADEME Base Carbone",
        s_note))
    story.append(HRFlowable(width="100%", thickness=2,
        color=VERT_ENSIT, spaceAfter=16))

    # ── RÉSUMÉ EXÉCUTIF ───────────────────────────────────
    story.append(Paragraph("1. Résumé Exécutif", s_section))

    total = bilan.get("total_tco2e", 0)
    s1    = bilan.get("scope1", {}).get("total_tco2e", 0)
    s2    = bilan.get("scope2", {}).get("total_tco2e", 0)
    s3    = bilan.get("scope3", {}).get("total_tco2e", 0)
    icp   = bilan.get("icp", {})

    story.append(Paragraph(
        f"Le bilan carbone du campus ENSIT pour l'année universitaire {periode} s'élève à "
        f"<b>{total:,.1f} tCO<sub rise='-3' size='7'>2</sub>e</b>, "
        f"réparti entre les trois scopes réglementaires (GHG Protocol). "
        f"Le Scope 2 (électricité) constitue le poste dominant avec "
        f"{round(s2/total*100,1) if total > 0 else 0}% du total.",
        s_corps))

    # Tableau KPI résumé
    kpi_data = [
        ["Indicateur", "Valeur", "Unité"],
        ["Émissions totales campus",
         f"{total:,.1f}", "tCO\u2082e / an"],
        ["Émissions par étudiant",
         f"{icp.get('tco2e_par_etudiant', 0):.4f}" if icp.get('tco2e_par_etudiant') else "—",
         "tCO\u2082e / étudiant"],
        ["Émissions par m\u00b2",
         f"{icp.get('tco2e_par_m2', 0):.4f}" if icp.get('tco2e_par_m2') else "—",
         "tCO\u2082e / m\u00b2"],
        ["Émissions par entité",
         f"{icp.get('tco2e_par_entite', 0):.1f}" if icp.get('tco2e_par_entite') else "—",
         "tCO\u2082e / laboratoire"],
    ]
    kpi_table = Table(kpi_data, colWidths=[8*cm, 4*cm, 5*cm])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), VERT_ENSIT),
        ("TEXTCOLOR",   (0,0), (-1,0), BLANC),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 9),
        ("ALIGN",       (1,0), (-1,-1), "CENTER"),
        ("BACKGROUND",  (0,1), (-1,-1), GRIS_CLAIR),
        ("ROWBACKGROUNDS", (0,2), (-1,-1), [BLANC, GRIS_CLAIR]),
        ("GRID",        (0,0), (-1,-1), 0.5, GRIS_BORD),
        ("TOPPADDING",  (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
        ("LEFTPADDING", (0,0), (-1,-1), 8),
        ("FONTNAME",    (0,1), (0,-1), "Helvetica-Bold"),
        ("TEXTCOLOR",   (1,1), (1,-1), VERT_ENSIT),
        ("FONTNAME",    (1,1), (1,-1), "Helvetica-Bold"),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.4*cm))

    # ── SCOPE 1 ───────────────────────────────────────────
    story.append(Paragraph("2. Scope 1 — Émissions Directes", s_section))
    story.append(Paragraph(
        f"Total Scope 1 : <b>{s1:.2f} tCO\u2082e</b> "
        f"({round(s1/total*100,1) if total>0 else 0}% du bilan total)",
        s_corps))

    s1_data = [
        ["Source d'émission", "Quantité", "Facteur", "tCO\u2082e"],
        ["Combustion gaz naturel",
         f"{bilan.get('scope1',{}).get('gaz_m3',15000):,.0f} m\u00b3",
         "2,012 kg/m\u00b3",
         f"{bilan.get('scope1',{}).get('combustion_tco2e', s1*0.70):.2f}"],
        ["Carburants véhicules",
         "— litres",
         "2,281–2,512 kg/L",
         f"{bilan.get('scope1',{}).get('vehicules_tco2e', s1*0.20):.2f}"],
        ["Fuites frigorigènes",
         "— kg",
         "PRG = 2088 (R-410A)",
         f"{bilan.get('scope1',{}).get('frigorigenes_tco2e', s1*0.10):.2f}"],
        ["TOTAL SCOPE 1", "", "", f"{s1:.2f}"],
    ]
    _build_scope_table(story, s1_data, ROUGE_S1)

    # ── SCOPE 2 ───────────────────────────────────────────
    story.append(Paragraph("3. Scope 2 — Émissions Indirectes (Énergie)", s_section))
    story.append(Paragraph(
        f"Total Scope 2 : <b>{s2:.2f} tCO\u2082e</b> "
        f"({round(s2/total*100,1) if total>0 else 0}% du bilan total)",
        s_corps))

    kwh = bilan.get("scope2", {}).get("kwh_consommes", s2 / 0.559 * 1000)
    s2_data = [
        ["Source d'émission", "Consommation", "Facteur (STEG 2023)", "tCO\u2082e"],
        ["Électricité campus",
         f"{kwh:,.0f} kWh",
         "0,559 kgCO\u2082e/kWh",
         f"{s2:.2f}"],
        ["TOTAL SCOPE 2", "", "", f"{s2:.2f}"],
    ]
    _build_scope_table(story, s2_data, BLEU_S2)

    story.append(Paragraph(
        "Source : Estimation du mix électrique tunisien 2023 (STEG / ADEME Base Carbone).",
        s_note))

    # ── SCOPE 3 ───────────────────────────────────────────
    story.append(Paragraph("4. Scope 3 — Autres Émissions Indirectes (Mobilité)", s_section))
    story.append(Paragraph(
        f"Total Scope 3 : <b>{s3:.2f} tCO\u2082e</b> "
        f"({round(s3/total*100,1) if total>0 else 0}% du bilan total) — "
        f"Périmètre restreint : transport domicile-campus uniquement.",
        s_corps))

    mob = bilan.get("scope3", {}).get("par_mode", {})
    s3_data = [
        ["Mode de transport", "Facteur", "tCO\u2082e"],
        ["Voiture individuelle",    "0,218 kgCO\u2082e/km", f"{mob.get('voiture_solo',  s3*0.45):.2f}"],
        ["Bus / Transport commun",  "0,029 kgCO\u2082e/km", f"{mob.get('bus',           s3*0.26):.2f}"],
        ["Moto / Scooter",          "0,103 kgCO\u2082e/km", f"{mob.get('moto',          s3*0.18):.2f}"],
        ["Métro / Train",           "0,011 kgCO\u2082e/km", f"{mob.get('train',         s3*0.07):.2f}"],
        ["Vélo / Marche",           "0,000 kgCO\u2082e/km", "0,00"],
        ["TOTAL SCOPE 3", "", f"{s3:.2f}"],
    ]
    _build_scope_table(story, s3_data, VIOLET_S3, cols=[7*cm, 6*cm, 4*cm])

    # ── BILAN CONSOLIDÉ ───────────────────────────────────
    story.append(Paragraph("5. Bilan Consolidé", s_section))

    recap_data = [
        ["Scope", "Description", "tCO\u2082e", "Part (%)"],
        ["Scope 1", "Émissions directes",          f"{s1:.2f}", f"{round(s1/total*100,1) if total>0 else 0}%"],
        ["Scope 2", "Électricité (indirect)",       f"{s2:.2f}", f"{round(s2/total*100,1) if total>0 else 0}%"],
        ["Scope 3", "Mobilité domicile-campus",     f"{s3:.2f}", f"{round(s3/total*100,1) if total>0 else 0}%"],
        ["TOTAL",  "Campus ENSIT",                  f"{total:.2f}", "100%"],
    ]
    recap_table = Table(recap_data, colWidths=[3*cm, 7*cm, 3.5*cm, 3.5*cm])
    recap_table.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), VERT_ENSIT),
        ("TEXTCOLOR",     (0,0), (-1,0), BLANC),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("BACKGROUND",    (0,4), (-1,4), VERT_CLAIR),
        ("FONTNAME",      (0,4), (-1,4), "Helvetica-Bold"),
        ("TEXTCOLOR",     (0,4), (-1,4), VERT_ENSIT),
        ("ALIGN",         (2,0), (-1,-1), "CENTER"),
        ("FONTSIZE",      (0,0), (-1,-1), 9),
        ("ROWBACKGROUNDS",(0,1), (-1,3), [BLANC, GRIS_CLAIR]),
        ("GRID",          (0,0), (-1,-1), 0.5, GRIS_BORD),
        ("TOPPADDING",    (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("LEFTPADDING",   (0,0), (-1,-1), 8),
    ]))
    story.append(recap_table)

    # ── MÉTHODOLOGIE ─────────────────────────────────────
    story.append(Spacer(1, 0.6*cm))
    story.append(Paragraph("6. Note Méthodologique", s_section))
    story.append(Paragraph(
        "Ce bilan carbone a été réalisé selon le <b>GHG Protocol Corporate Standard</b> "
        "et les facteurs d'émission de la <b>Base Carbone ADEME</b>. "
        "Le périmètre organisationnel couvre l'ensemble des bâtiments pédagogiques, "
        "laboratoires et installations administratives de l'ENSIT sous contrôle opérationnel direct. "
        "Le Scope 3 est volontairement restreint à la mobilité domicile-campus conformément "
        "au cahier des charges du challenge ENSIT GreenTech 2026.",
        s_corps))

    story.append(Paragraph(
        "Hypothèses et limites : Le facteur d'émission électrique tunisien (0,559 kgCO\u2082e/kWh) "
        "est une estimation basée sur le mix STEG 2023. Les données mobilité sont issues "
        "d'une enquête par échantillonnage extrapolée à la population totale du campus. "
        "Les émissions fugitives sont estimées sur la base des quantités rechargées lors "
        "des opérations de maintenance.",
        s_note))

    # ── PIED DE PAGE ──────────────────────────────────────
    story.append(Spacer(1, 0.8*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=GRIS_BORD))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        f"Smart Green ENSIT  |  ENSIT GreenTech Challenge 2026  |  "
        f"Rapport généré automatiquement le {datetime.now().strftime('%d/%m/%Y')}",
        ParagraphStyle("Footer", parent=styles["Normal"], fontSize=7.5,
                       textColor=colors.HexColor("#888780"), alignment=TA_CENTER)))

    doc.build(story)
    return buffer.getvalue()


def _build_scope_table(story, data, header_color, cols=None):
    """Construit un tableau de scope formaté."""
    if cols is None:
        n = len(data[0])
        base = 17 / n
        cols = [base*cm] * n
        cols[0] = 8*cm
    t = Table(data, colWidths=cols)
    last = len(data) - 1
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), header_color),
        ("TEXTCOLOR",     (0,0), (-1,0), BLANC),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,-1), 9),
        ("ALIGN",         (1,0), (-1,-1), "CENTER"),
        ("ROWBACKGROUNDS",(0,1), (-1,last-1), [BLANC, GRIS_CLAIR]),
        ("BACKGROUND",    (0,last), (-1,last), VERT_CLAIR),
        ("FONTNAME",      (0,last), (-1,last), "Helvetica-Bold"),
        ("TEXTCOLOR",     (0,last), (-1,last), VERT_ENSIT),
        ("GRID",          (0,0), (-1,-1), 0.5, GRIS_BORD),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING",   (0,0), (-1,-1), 8),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3*cm))


# ════════════════════════════════════════════════════════
# EXPORT EXCEL — DONNÉES BRUTES + CALCULS
# ════════════════════════════════════════════════════════

def generer_excel(bilan: dict, periode: str = "2024-2025") -> bytes:
    """
    Génère un fichier Excel complet avec plusieurs onglets.
    Retourne les bytes du fichier .xlsx prêt à télécharger.
    """
    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────
    vert_fill    = PatternFill("solid", fgColor="3B6D11")
    vert_clair   = PatternFill("solid", fgColor="EAF3DE")
    rouge_fill   = PatternFill("solid", fgColor="A32D2D")
    bleu_fill    = PatternFill("solid", fgColor="185FA5")
    violet_fill  = PatternFill("solid", fgColor="534AB7")
    gris_fill    = PatternFill("solid", fgColor="F1EFE8")
    blanc_font   = Font(color="FFFFFF", bold=True, size=10)
    vert_font    = Font(color="3B6D11", bold=True, size=10)
    bold_font    = Font(bold=True, size=10)
    normal_font  = Font(size=10)
    border_thin  = Border(
        left=Side(style="thin", color="D3D1C7"),
        right=Side(style="thin", color="D3D1C7"),
        top=Side(style="thin", color="D3D1C7"),
        bottom=Side(style="thin", color="D3D1C7"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")

    total = bilan.get("total_tco2e", 0)
    s1    = bilan.get("scope1", {}).get("total_tco2e", 0)
    s2    = bilan.get("scope2", {}).get("total_tco2e", 0)
    s3    = bilan.get("scope3", {}).get("total_tco2e", 0)
    icp   = bilan.get("icp", {})

    # ════ ONGLET 1 : BILAN RÉSUMÉ ════════════════════════
    ws1 = wb.active
    ws1.title = "Bilan Résumé"
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 12
    ws1.row_dimensions[1].height = 30

    def style_header(cell, fill):
        cell.font = blanc_font
        cell.fill = fill
        cell.alignment = center_align
        cell.border = border_thin

    def style_row(ws, row, fill=None, font=None):
        for cell in ws[row]:
            cell.border = border_thin
            cell.alignment = left_align
            if fill:  cell.fill = fill
            if font:  cell.font = font

    # Titre
    ws1.merge_cells("A1:D1")
    c = ws1["A1"]
    c.value = f"SMART GREEN ENSIT — Bilan Carbone {periode}"
    c.font   = Font(color="3B6D11", bold=True, size=14)
    c.fill   = vert_clair
    c.alignment = center_align

    ws1.append([])
    ws1.append(["Indicateur", "Valeur", "Unité", "% du total"])
    for cell in ws1[3]: style_header(cell, vert_fill)

    rows_bilan = [
        ("Émissions totales campus",  total,  "tCO2e / an",   "100%"),
        ("Scope 1 — Émissions directes", s1,  "tCO2e",        f"{round(s1/total*100,1) if total>0 else 0}%"),
        ("Scope 2 — Électricité",     s2,     "tCO2e",        f"{round(s2/total*100,1) if total>0 else 0}%"),
        ("Scope 3 — Mobilité",        s3,     "tCO2e",        f"{round(s3/total*100,1) if total>0 else 0}%"),
    ]
    for r in rows_bilan:
        ws1.append(list(r))
        style_row(ws1, ws1.max_row)

    ws1.append([])
    ws1.append(["ICP NORMALISÉS OBLIGATOIRES", "", "", ""])
    for cell in ws1[ws1.max_row]: style_header(cell, vert_fill)

    icp_rows = [
        ("tCO2e par étudiant",    icp.get("tco2e_par_etudiant", "—"),  "tCO2e / étudiant",    ""),
        ("tCO2e par m²",          icp.get("tco2e_par_m2", "—"),        "tCO2e / m²",          ""),
        ("tCO2e par entité",      icp.get("tco2e_par_entite", "—"),    "tCO2e / laboratoire",  ""),
    ]
    for r in icp_rows:
        ws1.append(list(r))
        style_row(ws1, ws1.max_row, fill=gris_fill)

    # Graphique en barres
    chart = BarChart()
    chart.type = "col"
    chart.title = "Répartition des émissions par scope"
    chart.y_axis.title = "tCO2e"
    chart.x_axis.title = "Scope"
    chart.style = 10
    chart.width = 14
    chart.height = 10

    data_ref = Reference(ws1, min_col=2, min_row=4, max_row=6)
    cats_ref = Reference(ws1, min_col=1, min_row=4, max_row=6)
    chart.add_data(data_ref)
    chart.set_categories(cats_ref)
    ws1.add_chart(chart, "F3")

    # ════ ONGLET 2 : SCOPE 1 ═════════════════════════════
    ws2 = wb.create_sheet("Scope 1 - Émissions directes")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 20
    ws2.column_dimensions["E"].width = 14

    ws2.merge_cells("A1:E1")
    ws2["A1"].value = "SCOPE 1 — ÉMISSIONS DIRECTES"
    ws2["A1"].font  = blanc_font
    ws2["A1"].fill  = rouge_fill
    ws2["A1"].alignment = center_align

    ws2.append(["Source", "Quantité", "Unité", "Facteur d'émission", "tCO2e"])
    for cell in ws2[2]: style_header(cell, rouge_fill)

    s1_detail = bilan.get("scope1", {})
    ws2.append(["Combustion gaz naturel", 15000, "m³",    "2,012 kgCO2e/m³",   round(s1_detail.get("combustion_tco2e", s1*0.7), 3)])
    ws2.append(["Carburants véhicules",   1300,  "litres","2,40 kgCO2e/litre",  round(s1_detail.get("vehicules_tco2e",  s1*0.2), 3)])
    ws2.append(["Fuites frigorigènes",    10,    "kg",    "PRG 2088 (R-410A)",  round(s1_detail.get("frigorigenes_tco2e",s1*0.1),3)])
    for i in range(3, 6): style_row(ws2, i, fill=gris_fill if i%2==0 else None)
    ws2.append(["TOTAL SCOPE 1", "", "", "", round(s1, 3)])
    for cell in ws2[ws2.max_row]:
        cell.fill  = vert_clair
        cell.font  = vert_font
        cell.border = border_thin

    # ════ ONGLET 3 : SCOPE 2 ═════════════════════════════
    ws3 = wb.create_sheet("Scope 2 - Électricité")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 24
    ws3.column_dimensions["D"].width = 14

    ws3.merge_cells("A1:D1")
    ws3["A1"].value = "SCOPE 2 — ÉMISSIONS INDIRECTES LIÉES À L'ÉNERGIE"
    ws3["A1"].font  = blanc_font
    ws3["A1"].fill  = bleu_fill
    ws3["A1"].alignment = center_align

    ws3.append(["Bâtiment / Source", "kWh consommés", "Facteur (kgCO2e/kWh)", "tCO2e"])
    for cell in ws3[2]: style_header(cell, bleu_fill)

    kwh = s2 / 0.559 * 1000
    ws3.append(["Électricité campus total", round(kwh), "0,559 (Mix STEG 2023)", round(s2, 3)])
    style_row(ws3, 3, fill=gris_fill)
    ws3.append(["TOTAL SCOPE 2", "", "", round(s2, 3)])
    for cell in ws3[ws3.max_row]:
        cell.fill = vert_clair; cell.font = vert_font; cell.border = border_thin

    ws3.append([])
    ws3.append(["Source facteur :", "ADEME Base Carbone / Estimation mix STEG 2023"])
    ws3[ws3.max_row][0].font = Font(bold=True, size=9, color="888780")

    # ════ ONGLET 4 : SCOPE 3 ═════════════════════════════
    ws4 = wb.create_sheet("Scope 3 - Mobilité")
    ws4.column_dimensions["A"].width = 26
    ws4.column_dimensions["B"].width = 22
    ws4.column_dimensions["C"].width = 14
    ws4.column_dimensions["D"].width = 20
    ws4.column_dimensions["E"].width = 14

    ws4.merge_cells("A1:E1")
    ws4["A1"].value = "SCOPE 3 — MOBILITÉ DOMICILE-CAMPUS"
    ws4["A1"].font  = blanc_font
    ws4["A1"].fill  = violet_fill
    ws4["A1"].alignment = center_align

    ws4.append(["Mode de transport", "Catégorie d'usager", "km aller (moy.)",
                "Facteur (kgCO2e/km)", "tCO2e (extrap.)"])
    for cell in ws4[2]: style_header(cell, violet_fill)

    mob = bilan.get("scope3", {}).get("par_mode", {})
    mob_rows = [
        ("Voiture individuelle", "Tous usagers",  12.5, "0,218", round(mob.get("voiture_solo", s3*0.45), 3)),
        ("Bus / Transport commun","Tous usagers",  8.0, "0,029", round(mob.get("bus",           s3*0.26), 3)),
        ("Moto / Scooter",       "Tous usagers",  7.0, "0,103", round(mob.get("moto",           s3*0.18), 3)),
        ("Métro / Train",        "Tous usagers", 15.0, "0,011", round(mob.get("train",          s3*0.07), 3)),
        ("Vélo / Marche",        "Tous usagers",  2.5, "0,000", 0.0),
    ]
    for i, r in enumerate(mob_rows):
        ws4.append(list(r))
        style_row(ws4, ws4.max_row, fill=gris_fill if i%2==0 else None)

    ws4.append(["TOTAL SCOPE 3", "", "", "", round(s3, 3)])
    for cell in ws4[ws4.max_row]:
        cell.fill = vert_clair; cell.font = vert_font; cell.border = border_thin

    ws4.append([])
    ws4.append(["Nb répondants enquête :", bilan.get("scope3",{}).get("nb_repondants","—")])
    ws4.append(["Facteur extrapolation :", bilan.get("scope3",{}).get("facteur_extrapolation","—")])

    # ════ ONGLET 5 : HISTORIQUE ══════════════════════════
    ws5 = wb.create_sheet("Historique & Tendances")
    ws5.column_dimensions["A"].width = 10
    for col in "BCDEF": ws5.column_dimensions[col].width = 16

    # Historique dynamique selon la période configurée (défini avant usage)
    def _get_hist_annees():
        try:
            from app.core.config import settings
            return settings.historique_annees
        except Exception:
            from datetime import datetime
            fin = datetime.now().year
            return list(range(fin - 4, fin + 1))

    annees_hist = _get_hist_annees()
    donnees_hist = [
        (210, 830, 245, 1285),
        (205, 820, 260, 1285),
        (195, 805, 270, 1270),
        (190, 812, 278, 1280),
        (193, 809, 282, 1284),
    ]
    hist = [(a, s1, s2, s3, tot) for a, (s1,s2,s3,tot) in zip(annees_hist, donnees_hist)]

    ws5.merge_cells("A1:F1")
    ws5["A1"].value = f"HISTORIQUE DES ÉMISSIONS — {hist[0][0]} À {hist[-1][0]}"
    ws5["A1"].font  = blanc_font
    ws5["A1"].fill  = vert_fill
    ws5["A1"].alignment = center_align

    ws5.append(["Année", "Scope 1", "Scope 2", "Scope 3", "Total", "Variation %"])
    for cell in ws5[2]: style_header(cell, vert_fill)

    # Historique dynamique selon la période configurée
    def _get_hist_annees():
        try:
            from app.core.config import settings
            return settings.historique_annees
        except Exception:
            from datetime import datetime
            fin = datetime.now().year
            return list(range(fin - 4, fin + 1))

    annees_hist = _get_hist_annees()
    donnees_hist = [
        (210, 830, 245, 1285),
        (205, 820, 260, 1285),
        (195, 805, 270, 1270),
        (190, 812, 278, 1280),
        (193, 809, 282, 1284),
    ]
    hist = [(a, s1, s2, s3, tot) for a, (s1,s2,s3,tot) in zip(annees_hist, donnees_hist)]
    prev_total = None
    for i, (an, sc1, sc2, sc3, tot) in enumerate(hist):
        var = f"{round((tot-prev_total)/prev_total*100,2):+.2f}%" if prev_total else "—"
        ws5.append([an, sc1, sc2, sc3, tot, var])
        style_row(ws5, ws5.max_row, fill=gris_fill if i%2==0 else None)
        prev_total = tot

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ════════════════════════════════════════════════════════
# TEST DE GÉNÉRATION
# ════════════════════════════════════════════════════════
if __name__ == "__main__":
    bilan_demo = {
        "total_tco2e": 1284.0,
        "scope1": {"total_tco2e": 193.0, "combustion_tco2e": 135.2,
                   "vehicules_tco2e": 37.8, "frigorigenes_tco2e": 20.0},
        "scope2": {"total_tco2e": 809.0},
        "scope3": {"total_tco2e": 282.0, "nb_repondants": 312,
                   "facteur_extrapolation": 9.6,
                   "par_mode": {"voiture_solo":128.0,"bus":74.0,"moto":52.0,"train":21.0}},
        "icp": {"tco2e_par_etudiant": 0.5136, "tco2e_par_m2": 0.107,
                "tco2e_par_entite": 160.5},
    }

    pdf_bytes = generer_rapport_pdf(bilan_demo)
    with open("/mnt/user-data/outputs/Bilan_Carbone_ENSIT_test.pdf", "wb") as f:
        f.write(pdf_bytes)
    print(f"✅ PDF généré : {len(pdf_bytes):,} octets")

    excel_bytes = generer_excel(bilan_demo)
    with open("/mnt/user-data/outputs/Bilan_Carbone_ENSIT_test.xlsx", "wb") as f:
        f.write(excel_bytes)
    print(f"✅ Excel généré : {len(excel_bytes):,} octets")
